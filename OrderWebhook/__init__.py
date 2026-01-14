import logging
import azure.functions as func
import os
import traceback
import json
from datetime import datetime

# Configure logging to ensure it captures properly in Azure
logger = logging.getLogger(__name__)

def main(req: func.HttpRequest) -> func.HttpResponse:
    try:
        # --- 1. CONFIGURATION GROUPING ---
        # We read all environment variables at the start of the function
        # This makes it easy to see what the function depends on
        BLOB_CONN_STR = os.environ.get("BLOB_CONN_STR")
        BLOB_CONTAINER = os.environ.get("BLOB_CONTAINER")
        EXCEL_BLOB_NAME = os.environ.get("EXCEL_BLOB_NAME")
        CATALOG_BLOB_NAME = os.environ.get("CATALOG_BLOB_NAME")

        OPENAI_ENDPOINT = os.environ.get("AZURE_OPENAI_ENDPOINT")
        OPENAI_KEY = os.environ.get("AZURE_OPENAI_KEY")
        WHISPER_DEPLOY = os.environ.get("AZURE_OPENAI_WHISPER_DEPLOYMENT")
        GPT_DEPLOY = os.environ.get("AZURE_OPENAI_GPT_DEPLOYMENT")

        TWILIO_SID = os.environ.get("TWILIO_SID")
        TWILIO_AUTH = os.environ.get("TWILIO_AUTH")
        TWILIO_NUMBER = os.environ.get("TWILIO_WHATSAPP_NUMBER")

        # Imports grouped for speed and error isolation
        from urllib.parse import parse_qs
        import uuid
        import tempfile

        logging.info("--- Processing New WhatsApp Request ---")

        # --- 2. VALIDATION ---
        if not all([BLOB_CONN_STR, OPENAI_KEY, TWILIO_SID]):
            logging.error("Missing critical environment variables. Check Azure App Settings.")
            return func.HttpResponse("Server configuration error", status_code=500)

        # --- 3. REQUEST PARSING ---
        content_type = (req.headers.get("Content-Type") or "").lower()
        if "application/x-www-form-urlencoded" in content_type:
            body = req.get_body().decode("utf-8")
            form = {k: v[0] for k, v in parse_qs(body).items()}
        else:
            form = req.get_json() if req.get_body() else {}

        media_url = form.get("MediaUrl0") or form.get("MediaUrl")
        from_number = form.get("From", "").replace("whatsapp:", "")

        if not media_url:
            logging.warning(f"Request from {from_number} ignored: No MediaUrl found.")
            return func.HttpResponse("Accepted", status_code=200)

        # --- 4. CORE LOGIC ---
        logging.info(f"Downloading audio for customer: {from_number}")
        voice_path = download_raw_voice(media_url, sid=TWILIO_SID, auth=TWILIO_AUTH)
        
        try:
            # Transcription (Whisper)
            logging.info("Sending audio to Azure OpenAI Whisper...")
            transcript = transcribe_whisper(voice_path, OPENAI_ENDPOINT, OPENAI_KEY, WHISPER_DEPLOY)
            logging.info(f"Transcription result: {transcript}")

            if not transcript.strip():
                logging.error("Whisper returned empty text.")
                return func.HttpResponse("Could not understand audio", status_code=200)

            # Catalog Context Loading
            logging.info("Loading product catalog for context...")
            catalog_content = get_catalog_context(BLOB_CONN_STR, BLOB_CONTAINER, CATALOG_BLOB_NAME)

            # AI matching and Extraction (GPT)
            # We pass the catalog directly into the prompt context
            logging.info("Extracting order details using catalog and GPT...")
            order_data = extract_order_with_pricing(
                transcript, 
                catalog_content, 
                OPENAI_ENDPOINT, 
                OPENAI_KEY, 
                GPT_DEPLOY
            )
            # Log the matches for debugging
            for item in order_data.get('items', []):
                status = "✅" if item.get('price_found') else "❌ NOT IN CATALOG"
                logging.info(f"{status} {item['name']} - {item.get('unit_price', 0)} AED")

            # Save to Order Log (Excel)
            logging.info(f"Updating Excel sheet: {EXCEL_BLOB_NAME}")
            log_to_excel(order_data, from_number, conn=BLOB_CONN_STR, container=BLOB_CONTAINER, blob=EXCEL_BLOB_NAME)

            # Send WhatsApp invoice to customer
            logging.info(f"Sending WhatsApp invoice to {from_number}")
            invoice_msg = format_invoice(order_data)
            send_whatsapp_message(from_number, invoice_msg, TWILIO_SID, TWILIO_AUTH, TWILIO_NUMBER)
            
            logging.info("--- Request Successfully Processed ---")
            return func.HttpResponse(json.dumps({"status": "success"}), mimetype="application/json")

        finally:
            # Always clean up the temp file even if the code crashes
            if os.path.exists(voice_path):
                os.remove(voice_path)
                logging.info("Temporary audio file deleted.")

    except Exception:
        # Log the full error for debugging in Application Insights
        logging.error(f"CRITICAL ERROR: {traceback.format_exc()}")
        return func.HttpResponse("Internal processing error", status_code=500)

# ---------------- HELPER FUNCTIONS (Encapsulated) ----------------

def download_raw_voice(url, sid, auth):
    import requests
    import tempfile
    import uuid
    r = requests.get(url, auth=(sid, auth), stream=True, timeout=30)
    r.raise_for_status()
    path = os.path.join(tempfile.gettempdir(), f"{uuid.uuid4().hex}.ogg")
    with open(path, "wb") as f:
        for chunk in r.iter_content(chunk_size=8192):
            f.write(chunk)
    return path

def transcribe_whisper(file_path, endpoint, key, deployment):
    from openai import AzureOpenAI
    client = AzureOpenAI(api_key=key, api_version="2024-06-01", azure_endpoint=endpoint)
    with open(file_path, "rb") as audio:
        result = client.audio.transcriptions.create(model=deployment, file=audio)
    return result.text

def extract_order_with_pricing(transcript, catalog, endpoint, key, deployment):
    """Uses GPT-4o-mini to match transcript against catalog with status flags."""
    from openai import AzureOpenAI
    client = AzureOpenAI(api_key=key, api_version="2024-02-15-preview", azure_endpoint=endpoint)
    
    logging.info(f"--- 🤖 GPT EXTRACTION STARTING ---")
    
    prompt = f"""
    CATALOG DATA (Pipe-Separated):
    {catalog}
    
    USER TRANSCRIPT: "{transcript}"
    
    TASK: Extract items and match them to the catalog.
    RULES:
    1. If the item matches a catalog SKU (even with minor typos), use the catalog price.
    2. If an item is NOT in the catalog (e.g. 'kitty banana'), set 'unit_price' to 0 and 'price_found' to false.
    3. Return valid JSON only.
    
    FORMAT:
    {{
      "items": [
        {{"name": "SKU Name", "qty": 1, "unit_price": 0.0, "total": 0.0, "price_found": true}}
      ],
      "currency": "AED"
    }}
    """
    
    try:
        response = client.chat.completions.create(
            model=deployment,
            messages=[{"role": "system", "content": prompt}],
            response_format={"type": "json_object"}
        )
        result = json.loads(response.choices[0].message.content)
        logging.info(f"✅ GPT matched {len(result.get('items', []))} items.")
        return result
    except Exception as e:
        logging.error(f"❌ GPT ERROR: {str(e)}")
        return {"items": []}

def format_invoice(data):
    """Formats a WhatsApp message. Marks missing catalog items as Out of Stock."""
    logging.info("Formatting WhatsApp invoice with Out of Stock logic...")
    
    currency = data.get('currency', 'AED')
    items = data.get('items', [])
    
    if not items:
        return "❌ *Order Error*\nWe couldn't recognize any items. Please try again."

    msg = "📝 *Order Summary*\n"
    msg += "---\n"
    
    grand_total = 0
    has_out_of_stock = False

    for item in items:
        name = item.get('name', 'Unknown Item')
        qty = item.get('qty', 1)
        price = item.get('unit_price', 0)
        found = item.get('price_found', True)
        
        # If item is in catalog and has a price > 0
        if found and price > 0:
            item_total = qty * price
            grand_total += item_total
            msg += f"• *{name}* (x{qty})\n  Price: {item_total} {currency}\n"
        else:
            # Item not in catalog (e.g., "Kitty Banana")
            msg += f"• ~{name}~ \n  ❌ *OUT OF STOCK*\n"
            has_out_of_stock = True

    msg += "---\n"
    msg += f"*Total Payable: {grand_total} {currency}*\n\n"
    
    if has_out_of_stock:
        msg += "⚠️ _Items crossed out above are currently unavailable and were not added to the total._\n\n"
        msg += "Would you like to replace the out-of-stock items with something else?"
    else:
        msg += "✅ All items are available! We are preparing your order now."
    return msg

def send_whatsapp_message(to, body, sid, auth, from_num):
    from twilio.rest import Client

    # 1. Clean and prefix the 'from' number
    sender = from_num.strip()
    if not sender.startswith("whatsapp:"):
        sender = f"whatsapp:{sender}"
    
    # 2. Clean and prefix the 'to' number
    recipient = to.strip()
    if not recipient.startswith("whatsapp:"):
        recipient = f"whatsapp:{recipient}"

    # --- LOGGING THE ATTEMPT ---
    logging.info("--- TWILIO OUTBOUND LOG ---")
    logging.info(f"SENDER:    [{sender}]")
    logging.info(f"RECIPIENT: [{recipient}]")
    logging.info(f"MESSAGE:   {body[:50]}...") # Logs first 50 chars of the message
    
    client = Client(sid, auth)
    
    try:
        message = client.messages.create(
            body=body,
            from_=sender,
            to=recipient
        )
        logging.info(f"SUCCESS: Message SID {message.sid}")
        return message.sid
    except Exception as e:
        # Logs the specific error from Twilio
        logging.error(f"TWILIO FAILURE: {str(e)}")
        raise e

def log_to_excel(data, customer, conn, container, blob):
    from azure.storage.blob import BlobServiceClient
    from openpyxl import load_workbook, Workbook
    import tempfile
    
    service = BlobServiceClient.from_connection_string(conn)
    b_client = service.get_blob_client(container, blob)
    tmp = os.path.join(tempfile.gettempdir(), f"sync_{customer}.xlsx")
    
    # Download existing or create new
    try:
        with open(tmp, "wb") as f: 
            f.write(b_client.download_blob().readall())
        wb = load_workbook(tmp)
    except Exception:
        wb = Workbook()
        wb.active.append(["Date", "Customer", "Items", "Total"])
    
    ws = wb.active
    # Create a string summary of items for the Excel cell
    summary = ", ".join([f"{i['name']} x{i['qty']}" for i in data.get('items', [])])
    
    ws.append([
        datetime.utcnow().strftime("%Y-%m-%d %H:%M"), 
        customer, 
        summary, 
        data.get('total')
    ])
    
    wb.save(tmp)
    with open(tmp, "rb") as f: 
        b_client.upload_blob(f, overwrite=True)

def get_catalog_context(conn_str, container, blob_name):
    import pandas as pd
    import pymupdf4llm
    import io
    from azure.storage.blob import BlobServiceClient
    logging.info(f"--- Loading Catalog: {blob_name} ---")
    
    try:
        service_client = BlobServiceClient.from_connection_string(conn_str)
        blob_client = service_client.get_blob_client(container=container, blob=blob_name)
        
        # Check if blob exists before downloading
        if not blob_client.exists():
            logging.error(f"❌ BLOB NOT FOUND: '{blob_name}' in container '{container}'")
            return "ERROR: Catalog file missing."

        blob_data = blob_client.download_blob().readall()
        ext = blob_name.split('.')[-1].lower()
        
        df = pd.DataFrame() # Initialize empty

        if ext == 'xlsb':
            # REQUIREMENT: pip install pyxlsb
            logging.info("Detected XLSB format. Using pyxlsb engine...")
            df = pd.read_excel(io.BytesIO(blob_data), engine='pyxlsb')
            
        elif ext in ['xlsx', 'xls']:
            logging.info(f"Detected {ext.upper()} format. Using default engine...")
            df = pd.read_excel(io.BytesIO(blob_data))
            
        elif ext == 'pdf':
            import pymupdf4llm
            logging.info("Detected PDF format. Converting to Markdown...")
            with open("/tmp/catalog.pdf", "wb") as f:
                f.write(blob_data)
            md_text = pymupdf4llm.to_markdown("/tmp/catalog.pdf")
            return md_text
        
        else:
            logging.warning(f"Unsupported format: {ext}")
            return "Warning: Unsupported catalog format."

        # --- DATA VALIDATION & CLEANING ---
        if df.empty:
            logging.error("❌ CATALOG LOADED BUT EMPTY: No rows found in the sheet.")
            return "No catalog data available."

        # CLEANING: Remove empty rows/cols and strip whitespace
        df.dropna(how='all', inplace=True)
        df.columns = [str(c).strip() for c in df.columns]
        
        # Log success details
        logging.info(f"✅ SUCCESS: Catalog loaded with {len(df)} rows.")
        logging.info(f"Columns available: {list(df.columns)}")

        # Convert to Pipe-Separated CSV (More token-efficient for GPT than JSON or standard CSV)
        return df.to_csv(index=False, sep="|")

    except Exception as e:
        logging.error(f"❌ CRITICAL CATALOG ERROR: {str(e)}", exc_info=True)
        return "No catalog data available due to a system error."